import argparse
import json
import sys
from pathlib import Path

import openpyxl.cell
import openpyxl.worksheet.worksheet

from .constants import STATE_CODE_MAP


def get_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "src_xlsx",
        metavar="<EXCEL PATH>",
        help="Path to the source Tax foundation excel spreadsheet",
    )
    parser.add_argument(
        "dest_json", metavar="<JSON PATH>", help="Path to save the output JSON file"
    )

    args = parser.parse_args()
    return args


def get_state_code(value: str) -> str:
    if not value:
        return ""
    safe_value = value.lower().replace(".", "").split("(")[0]
    return STATE_CODE_MAP.get(safe_value, "")


def clean_deduction(value: str) -> tuple[str, bool]:
    """Return clean value, and True if it was a credit"""

    if isinstance(value, int):
        return {"value": value, "credit": False}
    if not value or value.lower().strip().replace(".", "") == "na":
        return {"value": None, "credit": False}
    if "credit" in value.lower():
        value = value.replace("credit", "").replace("$", "").replace(",", "").strip()
        value = int(value)
        return {"value": value, "credit": True}
    return {"value": value, "credit": False}


def extract_deductions(row: tuple[openpyxl.cell.cell.Cell]):
    standard_deductions = {
        "single": clean_deduction(row[7].value),
        "married": clean_deduction(row[8].value),
    }
    personal_exemptions = {
        "single": clean_deduction(row[9].value),
        "married": clean_deduction(row[10].value),
        "dependant": clean_deduction(row[11].value),
    }
    return {
        "standard_deductions": standard_deductions,
        "personal_exemptions": personal_exemptions,
    }


def extract_codes(row: openpyxl.cell.cell.Cell) -> list[str]:
    if isinstance(row[0].value, str) and ("(" in row[0].value):
        code_str = row[0].value.split("(")[1].replace(")", "")
        codes = [code.strip() for code in code_str.split(",")]
        return codes
    return []


def extract_notes(
    row1: tuple[openpyxl.cell.cell.Cell], row2: tuple[openpyxl.cell.cell.Cell]
) -> list[str]:
    notes = []
    note_codes = []

    for row in (row1, row2):
        for cell in row[12:]:
            if cell.value:
                notes.append(cell.value)
        note_codes += extract_codes(row)

    return notes, note_codes


def process_state(rows: list[tuple[openpyxl.cell.cell.Cell]]) -> dict:
    state_code = get_state_code(rows[0][0].value)

    deductions = extract_deductions(rows[0])
    notes, note_codes = extract_notes(rows[0], rows[1] if len(rows) > 1 else None)

    if isinstance(rows[0][2], openpyxl.cell.cell.MergedCell):
        # Don't have rates, just 'none' or some special case.
        notes.append(rows[0][1].value)
        brackets = {
            "single": [],
            "married": [],
            "notes": notes,
            "note_codes": note_codes,
            "deductions": deductions,
        }

    else:
        single = []
        married = []
        for row in rows:
            key_cols = (1, 3, 4, 6)
            if any([x.value for idx, x in enumerate(row) if idx in key_cols]):
                single.append({"rate": row[1].value, "start_value": row[3].value})
                married.append({"rate": row[4].value, "start_value": row[6].value})
        brackets = {
            "single": single,
            "married": married,
            "note_codes": note_codes,
            "notes": notes,
            "deductions": deductions,
        }

    return state_code, brackets


def process_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, dict]:
    """
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): Worksheet containing the data for a given year.

    Returns:
        dict[str,dict]: a dict with keys equal to two letter state abbreviations and values
    dicts containing state level data.
    """

    states = {}
    state_rows = []
    for row in sheet.iter_rows():
        # row is a tuple of openpyxl.cell.cell.Cell
        if isinstance(row[1], openpyxl.cell.cell.MergedCell):
            # We're in the notes section
            break
        colA = row[0]
        state_code = get_state_code(colA.value)
        if state_code:
            if state_rows:
                state, data = process_state(state_rows)
                states[state] = data
            state_rows = [row]
        else:
            state_rows.append(row)
    print(json.dumps(states["MD"]))

    return states


def write_json(dst_json, results: dict):
    path = Path(dst_json)
    path.write_text(json.dumps(results, indent=4), encoding="utf8")


def main():
    args = get_args()
    wb = openpyxl.load_workbook(args.src_xlsx)
    print("File read, processing...")
    names = [name for name in wb.sheetnames if len(name) == 4 and name.startswith("20")]
    names.sort(reverse=True)
    if not names:
        print("No appropriately named worksheets found (expected to be years)")
        print(f"Worksheets: {wb.worksheets}")
        sys.exit(1)

    results = {}
    for name in names:
        print(f"Processing {name}")
        sheet = wb[name]
        results[name] = process_sheet(sheet)
        # json.dumps(results[name]["MD"]["single"], indent=4)
        break

    write_json(args.dest_json, results)


if __name__ == "__main__":
    main()
    main()
